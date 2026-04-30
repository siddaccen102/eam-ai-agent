"""
generate_tracker.py — builds an Excel tracker for the EAM AI Agent project.

Run from project root:
    python tools/generate_tracker.py

Outputs:
    docs/project-tracker.xlsx

Two sheets:
  - "Scrims"        full row-per-quest breakdown with status colors, XP, weight, commit, notes
  - "Phase Summary" rollup by phase with completion %

When you finish a scrim, edit the matching Scrim(...) row below and re-run.
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).resolve().parent.parent / "docs" / "project-tracker.xlsx"

# ---------------------------------------------------------------------------
# Data: one row per scrim / level / mini-quest
# ---------------------------------------------------------------------------

@dataclass
class Scrim:
    id: str
    name: str
    phase: str
    kind: str          # Foundation | Level | Mini-quest
    status: str        # Completed | In Progress | Pending
    weight: float      # rough % of total project
    xp: int            # XP earned (0 if not yet started)
    commit: str = ""   # short commit hash if landed
    notes: str = ""

SCRIMS: list[Scrim] = [
    # ---- Foundations -------------------------------------------------------
    Scrim("F1",  "Repo bootstrap + monorepo structure",            "Foundations",     "Foundation",  "Completed",   4, 20, "0e4954c", "Docs + workspace layout"),
    Scrim("F2",  "Backend Express + TS foundation",                "Foundations",     "Foundation",  "Completed",   5, 25, "2940927", "Express 5 + tsx watch"),
    Scrim("F3",  "Frontend Vite + React + TS + Tailwind",          "Foundations",     "Foundation",  "Completed",   5, 25, "4510707", "Tailwind, Vite config"),
    Scrim("F4",  "Integration contracts baseline (docs)",          "Foundations",     "Foundation",  "Completed",   4, 20, "fa93827", "ARCHITECTURE.md, contracts.md"),
    Scrim("F5",  "Typed env validation (Zod)",                     "Foundations",     "Foundation",  "Completed",   5, 25, "533b6b7", "Fail-loud at startup"),

    # ---- Workday integration ----------------------------------------------
    Scrim("L1",   "IntegrationError contract module",              "Workday",         "Level",       "Completed",   3, 25, "8ec3e06", "Stable error codes + correlationId"),
    Scrim("L2",   "Workday Bridge - axios adapter + interceptors", "Workday",         "Level",       "Completed",   4, 64, "bec48c6", "Bonus: request logger"),
    Scrim("L2.5", "Workday smoke route + getWorkday helper",       "Workday",         "Level",       "Completed",   3, 30, "",        "Thin client helper pattern"),
    Scrim("L3",   "Phantom Workday - mock layer + WORKDAY_MODE",   "Workday",         "Level",       "Completed",   2, 20, "d4527dc", "Env flag committed; mock layer follow-up"),
    Scrim("L4",   "Workday -> ValidatedUser mapping",              "Workday",         "Level",       "Pending",     3,  0, "",        "Blocked on real Workday creds"),

    # ---- EAM integration ---------------------------------------------------
    Scrim("M4.1", "EAM Bridge - axios adapter (mirror of Workday)","EAM",             "Mini-quest",  "Completed",   4, 51, "",        "Bonus: duration logging"),
    Scrim("M4.2", "EAM helpers (getEam + postEam)",                "EAM",             "Mini-quest",  "Completed",   2, 23, "",        "Two-generic POST pattern"),
    Scrim("M4.4", "Real auth wiring (Basic + tenant/role/org)",    "EAM",             "Mini-quest",  "Completed",   3, 30, "a2ff977", "Bonus: masked startup log"),
    Scrim("M4.5", "First Real Call - assets smoke route",          "EAM",             "Mini-quest",  "Completed",   3, 39, "ed1c3f6", "Live Vopak data verified"),
    Scrim("M4.6", "Unwrap the envelope (Result.ResultData)",       "EAM",             "Mini-quest",  "In Progress", 2,  0, "",        "Anti-corruption layer pattern"),
    Scrim("M4.7", "Map to canonical EquipmentOption DTO",          "EAM",             "Mini-quest",  "Pending",     3,  0, "",        "First real DTO mapping"),
    Scrim("M5.1", "Sites/Departments mapping (terminal lookup)",   "EAM",             "Mini-quest",  "Pending",     3,  0, "",        "OrganizationContext DTO"),
    Scrim("M5.2", "Problem codes mapping",                         "EAM",             "Mini-quest",  "Pending",     2,  0, "",        "ProblemCodeOption DTO"),
    Scrim("M5.3", "Work-order POST + create flow",                 "EAM",             "Mini-quest",  "Pending",     5,  0, "",        "First real POST integration"),

    # ---- Business logic / orchestration -----------------------------------
    Scrim("L5",   "Backend business route handlers (5 endpoints)", "Business Logic",  "Level",       "Pending",     6,  0, "",        "Replaces /smoke routes"),
    Scrim("L6",   "Workflow orchestration service (agent core)",   "Business Logic",  "Level",       "Pending",     8,  0, "",        "Coordinates Workday + EAM steps"),

    # ---- Frontend ----------------------------------------------------------
    Scrim("L7",   "Frontend state mgmt + workflow UI",             "Frontend",        "Level",       "Pending",    11,  0, "",        "Wire each workflow step"),

    # ---- AI guidance layer -------------------------------------------------
    Scrim("L8",   "AI guidance layer (OpenAI integration)",        "AI",              "Level",       "Pending",     5,  0, "",        "Guidance only, not source of truth"),

    # ---- Quality & ship ---------------------------------------------------
    Scrim("L9",   "Tests + e2e validation",                        "Tests",           "Level",       "Pending",     2,  0, "",        "Smoke + integration"),
    Scrim("L10",  "Deployment + docs polish",                      "Deploy",          "Level",       "Pending",     2,  0, "",        "Hosting, README, runbook"),
]

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

STATUS_FILL = {
    "Completed":   PatternFill("solid", fgColor="C6EFCE"),  # green
    "In Progress": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "Pending":     PatternFill("solid", fgColor="EAEAEA"),  # grey
}
STATUS_FONT = {
    "Completed":   Font(color="006100", bold=True),
    "In Progress": Font(color="9C5700", bold=True),
    "Pending":     Font(color="595959"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT  = Font(bold=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ["ID", "Name", "Phase", "Kind", "Status", "Weight %", "Earned %", "XP", "Commit", "Notes"]
COL_WIDTHS = [8, 50, 18, 13, 14, 11, 11, 8, 12, 50]


def earned_for(scrim: Scrim) -> float:
    if scrim.status == "Completed":
        return scrim.weight
    if scrim.status == "In Progress":
        return scrim.weight * 0.5
    return 0.0


def build_scrims_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Scrims"

    # Header row
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for col, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    total_weight = sum(s.weight for s in SCRIMS)
    total_earned_pct = 0.0
    total_xp = 0

    for r, s in enumerate(SCRIMS, start=2):
        earned_pct = earned_for(s)
        total_earned_pct += earned_pct
        total_xp += s.xp

        row_values = [
            s.id, s.name, s.phase, s.kind, s.status,
            s.weight, round(earned_pct, 2), s.xp, s.commit, s.notes,
        ]
        for col, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 10))

        status_cell = ws.cell(row=r, column=5)
        status_cell.fill = STATUS_FILL[s.status]
        status_cell.font = STATUS_FONT[s.status]
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row = len(SCRIMS) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=6, value=total_weight)
    ws.cell(row=total_row, column=7, value=round(total_earned_pct, 2))
    ws.cell(row=total_row, column=8, value=total_xp)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER

    ws.freeze_panes = "A2"


def build_phase_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase Summary")

    headers = ["Phase", "Total Weight %", "Earned %", "Status", "# Items", "# Done"]
    widths = [22, 16, 14, 18, 10, 10]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    by_phase: dict[str, list[Scrim]] = {}
    for s in SCRIMS:
        by_phase.setdefault(s.phase, []).append(s)

    row = 2
    for phase, items in by_phase.items():
        weight = sum(i.weight for i in items)
        earned = sum(earned_for(i) for i in items)
        done = sum(1 for i in items if i.status == "Completed")
        completion_pct = (earned / weight * 100) if weight else 0
        status_label = "Done" if completion_pct >= 99 else ("In Progress" if completion_pct > 0 else "Not Started")

        for col, val in enumerate([phase, weight, round(earned, 2), status_label, len(items), done], start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if col != 1 else "left")
        sc = ws.cell(row=row, column=4)
        if status_label == "Done":
            sc.fill, sc.font = STATUS_FILL["Completed"], STATUS_FONT["Completed"]
        elif status_label == "In Progress":
            sc.fill, sc.font = STATUS_FILL["In Progress"], STATUS_FONT["In Progress"]
        else:
            sc.fill, sc.font = STATUS_FILL["Pending"], STATUS_FONT["Pending"]
        row += 1

    weight = sum(s.weight for s in SCRIMS)
    earned = sum(earned_for(s) for s in SCRIMS)
    done = sum(1 for s in SCRIMS if s.status == "Completed")
    pct = round(earned / weight * 100, 1) if weight else 0
    for col, val in enumerate(["PROJECT TOTAL", weight, round(earned, 2), f"{pct}%", len(SCRIMS), done], start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill, cell.font = TOTAL_FILL, TOTAL_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if col != 1 else "left")

    ws.freeze_panes = "A2"


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_scrims_sheet(wb)
    build_phase_sheet(wb)
    wb.save(OUTPUT_PATH)

    completed = sum(1 for s in SCRIMS if s.status == "Completed")
    earned = sum(earned_for(s) for s in SCRIMS)
    weight = sum(s.weight for s in SCRIMS)
    xp = sum(s.xp for s in SCRIMS)
    pct = round(earned / weight * 100, 1) if weight else 0

    print(f"Wrote {OUTPUT_PATH}")
    print(f"  {completed}/{len(SCRIMS)} scrims complete  |  {pct}% earned  |  {xp} total XP")


if __name__ == "__main__":
    main()
